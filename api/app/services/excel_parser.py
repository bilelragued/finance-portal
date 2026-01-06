"""Excel file parser for bank transaction exports."""
import re
from datetime import datetime
from typing import List, Dict, Tuple, Optional
from openpyxl import load_workbook
from pathlib import Path


class ExcelParser:
    """Parse bank transaction Excel files."""
    
    # Expected column headers (NZ bank format)
    EXPECTED_HEADERS = [
        "Transaction Date", "Processed Date", "Type", "Details", 
        "Particulars", "Code", "Reference", "Amount", "Balance",
        "To/From Account Number", "Conversion Charge", "Foreign Currency Amount"
    ]
    
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self.sheet = None
        
    def parse_filename(self) -> Dict:
        """
        Extract account number and date range from the file path's filename.
        """
        return self.parse_filename_string(self.file_path.name)
    
    def parse_filename_string(self, filename: str) -> Dict:
        """
        Extract account number and date range from a filename string.
        
        Expected format: 01-0183-0950462-00_Transactions_2025-06-01_2025-11-30.xlsx
        """
        # Remove extension if present
        if '.' in filename:
            filename_stem = filename.rsplit('.', 1)[0]
        else:
            filename_stem = filename
        
        result = {
            "account_number": None,
            "date_from": None,
            "date_to": None,
            "raw_filename": filename
        }
        
        # Pattern to match: account_number_Transactions_date-from_date-to
        pattern = r'^([\d-]+)_Transactions_(\d{4}-\d{2}-\d{2})_(\d{4}-\d{2}-\d{2})$'
        match = re.match(pattern, filename_stem)
        
        if match:
            result["account_number"] = match.group(1)
            result["date_from"] = datetime.strptime(match.group(2), "%Y-%m-%d").date()
            result["date_to"] = datetime.strptime(match.group(3), "%Y-%m-%d").date()
        else:
            # Try to extract just account number (any format starting with digits and dashes)
            account_pattern = r'^([\d]+-[\d]+-[\d]+-[\d]+)'
            account_match = re.match(account_pattern, filename_stem)
            if account_match:
                result["account_number"] = account_match.group(1)
        
        return result
    
    def load(self) -> bool:
        """Load the Excel file."""
        try:
            self.workbook = load_workbook(self.file_path, read_only=True, data_only=True)
            # Get the first sheet (usually named "Transactions")
            self.sheet = self.workbook.active
            return True
        except Exception as e:
            raise ValueError(f"Failed to load Excel file: {e}")
    
    def validate_headers(self) -> Tuple[bool, List[str]]:
        """Validate that the file has expected column headers."""
        if not self.sheet:
            raise ValueError("File not loaded. Call load() first.")
        
        # Read first row as headers
        headers = [cell.value for cell in self.sheet[1]]
        
        missing = []
        for expected in self.EXPECTED_HEADERS[:9]:  # First 9 are required
            if expected not in headers:
                missing.append(expected)
        
        return len(missing) == 0, missing
    
    def parse_amount(self, amount_str: str) -> float:
        """
        Parse amount string to float.
        
        Handles formats like:
        - "- $123.45" (debit)
        - "$123.45" (credit)
        - "-$123.45"
        """
        if not amount_str or amount_str == "":
            return 0.0
        
        # Convert to string if not already
        amount_str = str(amount_str).strip()
        
        # Remove currency symbol and spaces
        cleaned = amount_str.replace("$", "").replace(",", "").replace(" ", "")
        
        # Handle "- " prefix for debits
        if cleaned.startswith("-"):
            cleaned = cleaned[1:]
            return -float(cleaned)
        
        return float(cleaned)
    
    def parse_date(self, date_value) -> Optional[datetime]:
        """Parse date from Excel cell value."""
        if not date_value:
            return None
        
        # If already a datetime object
        if isinstance(date_value, datetime):
            return date_value.date()
        
        # Try common date formats
        date_str = str(date_value).strip()
        formats = [
            "%d %b %Y",      # "28 Nov 2025"
            "%Y-%m-%d",      # "2025-11-28"
            "%d/%m/%Y",      # "28/11/2025"
            "%d-%m-%Y",      # "28-11-2025"
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        
        return None
    
    def extract_card_last4(self, details: str) -> Optional[str]:
        """Extract last 4 digits of card from details field."""
        if not details:
            return None
        
        # Pattern: 4835-****-****-3704
        pattern = r'\d{4}-\*{4}-\*{4}-(\d{4})'
        match = re.search(pattern, details)
        
        if match:
            return match.group(1)
        return None
    
    def parse_transactions(self) -> List[Dict]:
        """
        Parse all transactions from the Excel file.
        
        Returns list of transaction dictionaries.
        """
        if not self.sheet:
            raise ValueError("File not loaded. Call load() first.")
        
        transactions = []
        
        # Get header row to map columns
        headers = [cell.value for cell in self.sheet[1]]
        header_map = {header: idx for idx, header in enumerate(headers) if header}
        
        # Parse each row (skip header)
        for row_num, row in enumerate(self.sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not any(row):
                continue
            
            def get_value(column_name):
                idx = header_map.get(column_name)
                if idx is not None and idx < len(row):
                    return row[idx]
                return None
            
            # Parse transaction date
            trans_date = self.parse_date(get_value("Transaction Date"))
            if not trans_date:
                continue  # Skip rows without valid date
            
            # Parse amount
            amount_raw = get_value("Amount")
            try:
                amount = self.parse_amount(amount_raw)
            except (ValueError, TypeError):
                continue  # Skip rows with invalid amount
            
            # Parse balance
            balance_raw = get_value("Balance")
            try:
                balance = self.parse_amount(balance_raw) if balance_raw else None
            except (ValueError, TypeError):
                balance = None
            
            # Extract details
            details = get_value("Details")
            
            transaction = {
                "transaction_date": trans_date,
                "processed_date": self.parse_date(get_value("Processed Date")),
                "transaction_type": get_value("Type"),
                "details": details,
                "particulars": get_value("Particulars"),
                "code": get_value("Code"),
                "reference": get_value("Reference"),
                "amount": amount,
                "balance": balance,
                "to_from_account": get_value("To/From Account Number"),
                "conversion_charge": get_value("Conversion Charge"),
                "foreign_currency_amount": get_value("Foreign Currency Amount"),
                "card_number_last4": self.extract_card_last4(str(details) if details else ""),
                "row_number": row_num  # For debugging/reference
            }
            
            transactions.append(transaction)
        
        return transactions
    
    def get_summary(self) -> Dict:
        """Get summary information about the file."""
        file_info = self.parse_filename()
        
        if not self.sheet:
            self.load()
        
        transactions = self.parse_transactions()
        
        if not transactions:
            return {
                **file_info,
                "total_rows": 0,
                "date_range": None,
                "opening_balance": None,
                "closing_balance": None,
            }
        
        # Transactions are typically in reverse chronological order (newest first)
        # Sort by date to get proper range
        sorted_trans = sorted(transactions, key=lambda x: x["transaction_date"])
        
        return {
            **file_info,
            "total_rows": len(transactions),
            "date_range": {
                "from": sorted_trans[0]["transaction_date"],
                "to": sorted_trans[-1]["transaction_date"]
            },
            # Opening balance is the balance AFTER the oldest transaction
            # Closing balance is the balance AFTER the newest transaction
            "oldest_transaction": sorted_trans[0],
            "newest_transaction": sorted_trans[-1],
            "opening_balance": sorted_trans[0]["balance"],
            "closing_balance": sorted_trans[-1]["balance"],
        }
    
    def close(self):
        """Close the workbook."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.sheet = None


def parse_excel_file(file_path: str) -> Tuple[Dict, List[Dict]]:
    """
    Convenience function to parse an Excel file.
    
    Returns (file_info, transactions)
    """
    parser = ExcelParser(file_path)
    try:
        parser.load()
        valid, missing = parser.validate_headers()
        if not valid:
            raise ValueError(f"Missing required columns: {missing}")
        
        summary = parser.get_summary()
        transactions = parser.parse_transactions()
        
        return summary, transactions
    finally:
        parser.close()

